#!/usr/bin/python
# Filename: cIO.py
try:
    import os, sys, logging
    sys.path.append(os.path.join(os.path.dirname(__file__), "openpyxl"))
    import openpyxl as oxl
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging, openpyxl).")


class Read:
    def __init__(self, *args):
        # args[0] = full_file_name --  absolute path of a workbook
        # args[1] = worksheet -- INT sheet number in workbook
        # args[2] = logger

        try:
            self.open_wb(args[0], True, True)  # open workbook in read_only and data_only modes
        except:
            self.wb = ""

        try:
            self.open_ws(int(args[1]))
        except:
            self.ws = ""

        try:
            self.logger = logging.getLogger(args[2])
        except:
            self.logger = logging.getLogger("logfile.log")

    def close_wb(self):
        ###
        self.wb.close()

    def open_wb(self, xlsx_name, read_only=True, data_only=True):
        ###
        wb = oxl.load_workbook(filename=xlsx_name, read_only=read_only, data_only=data_only)
        self.wb = wb

    def open_ws(self, worksheet):
        # worksheet = INT
        ws = self.wb.worksheets[worksheet]
        self.ws = ws

    def read_column(self, column, start_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = STR, e.g., col = "B"
        # start_row = INT
        # returns column as LIST

        data = []
        test_content = True
        row = start_row
        while test_content:
            code = column + str(start_row)
            value = self.ws[code].value
            if value is None:
                test_content = False
            else:
                data.append(value)
        
        return data

    def __call__(self):
        print("Class Info: <type> = XLSX manipulation in cIO.py")


class WbMod(Read):
    def __init__(self, *args):
        # args[0] = STR full_path to workbook template
        # args[1] = INT worksheet number
        # args[2] = STR logger
        try:
            Read.__init__(self, "#", "#", args[1])
        except:
            Read.__init__(self)

        try:
            self.open_wb(args[0], read_only=False)
            try:
                self.open_ws(args[1])
            except:
                pass
        except:
            pass


    def save_close_wb(self, full_file_path):
        ###
        self.wb.save(full_file_path)
        self.wb.close()

    def write_data_cell(self, column, row, value):
        ###
        code = column + str(row)
        self.ws[code].value = value


    def write_data_column(self, column, start_row, data_list):
        # writes COLUMN beginning at START_ROW
        # data_list is a LIST object
        self.logger.info("   * Writing column data (starting at " + str(column) + str(start_row) + ") ...")

        row = start_row
        for datum in data_list:
            code = column + str(row)
            self.ws[code].value = datum
            row += 1

